"""
climbro.render.sheets_static — Mobility, Nutrition, Recovery, Glossary.

Mobility and Nutrition are flag-gated by the config; Recovery and Glossary are
always included. English, concise. Nutrition targets read the latest bodyweight
from the Week sheet (falling back to the start weight).
"""

from __future__ import annotations
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment
from .style import h1, h2, th, td, note, fill, inp, widths, FONT, GREY, RED, GREEN, YELLOW, ORANGE, LBLUE
from ..engine import Plan
from ..i18n import translator


# --------------------------------------------------------------------------- #
# Mobility (optional)
# --------------------------------------------------------------------------- #
def build_mobility(ws, plan: Plan) -> None:
    t = translator(plan.cfg.language)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1"); h1(ws["A1"], t("Mobility — protocol, measurements, periodization"))
    ws.row_dimensions[1].height = 24
    widths(ws, {"A": 20, "B": 16, "C": 30, "D": 12, "E": 12, "F": 12, "G": 12, "H": 18})

    h2(ws, 3, 1, 8, t("Protocol (climbing-specific: active mobility, not passive splits)"))
    th(ws.cell(4, 1, t("When"))); th(ws.cell(4, 2, t("Frequency")))
    ws.merge_cells("C4:F4"); th(ws.cell(4, 3, t("What")))
    ws.merge_cells("G4:H4"); th(ws.cell(4, 7, t("Why")))
    rows = [
        (t("Dynamic warm-up"), t("Every session"), t("Leg/hip swings, 90-90, glute bridge, band pull-aparts & dislocates, cat-cow, wrists (8-10')"), t("Prep tissue. No static stretch pre-session — it cuts force.")),
        (t("Targeted mobility"), t("2x/week"), t("Hips (90-90, deep squat hold, frog, couch), shoulders/T-spine (hangs, rotations), ankles (knee-to-wall). Active holds (20-25')"), t("Usable range for climbing positions")),
        (t("PNF / contract-relax"), t("1x/week"), t("Enter stretch → 5-6s ~70% contraction → relax deeper, 3-4 cycles. Stubborn areas (hips)"), t("Best range gains for strength-based athletes")),
    ]
    r = 5
    for a, b, c, d in rows:
        td(ws.cell(r, 1), a, bold=True); td(ws.cell(r, 2), b, center=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6); td(ws.cell(r, 3), c)
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8); td(ws.cell(r, 7), d)
        ws.row_dimensions[r].height = 44
        r += 1

    r += 1
    h2(ws, r, 1, 8, t("Measurements (every 2-4 weeks — range changes slowly)")); ws.row_dimensions[r].height = 18
    r += 1
    headers = [t("Date"), t("Deep squat hold, s"), t("Sit-and-reach, cm"), t("Shoulder wall 0/1/2"),
               t("Dorsi L, cm"), t("Dorsi R, cm"), t("Asym |L−R|"), t("Tightness 0-3")]
    for i, htext in enumerate(headers, 1):
        th(ws.cell(r, i, htext))
    ws.row_dimensions[r].height = 44
    first = r + 1
    for k in range(14):
        rr = first + k
        for c in (1, 2, 3, 4, 5, 6, 8):
            td(ws.cell(rr, c), center=True); fill(ws.cell(rr, c), YELLOW)
        ws.cell(rr, 1).number_format = "dd.mm.yyyy"
        ws.cell(rr, 7, f'=IF(OR(E{rr}="",F{rr}=""),"",ABS(E{rr}-F{rr}))'); ws.cell(rr, 7).alignment = Alignment("center", "center")
        ws.row_dimensions[rr].height = 20
    ws.conditional_formatting.add(f"G{first}:G{first+13}", CellIsRule(operator="greaterThan", formula=["2"], fill=PatternFill("solid", fgColor=RED)))
    note(ws.cell(first + 15, 1),
         t("Yellow = your input. Asymmetry auto-computes (>2 cm highlighted — worth balancing). "
           "Shoulder wall: 0 = can't reach overhead without arching, 1 = partial, 2 = easy. Same conditions each time."))
    ws.merge_cells(start_row=first + 15, start_column=1, end_row=first + 15, end_column=8); ws.row_dimensions[first + 15].height = 38
    rr = first + 17
    h2(ws, rr, 1, 8, t("Periodization (in sync with the main cycle)")); rr += 1
    for a, b in [(t("Base / contact (deficit)"), t("Most mobility volume — best window to build range. 2 targeted + 1 PNF per week.")),
                 (t("Bridge / peak"), t("Reduce deep stretching and PNF (they add fatigue before the peak). Dynamic + light maintenance only.")),
                 (t("Taper"), t("Dynamic warm-up and light mobilization only. Nothing new or intense."))]:
        td(ws.cell(rr, 1), a, bold=True); ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=8); td(ws.cell(rr, 2), b)
        ws.row_dimensions[rr].height = 30; rr += 1


# --------------------------------------------------------------------------- #
# Nutrition (optional)
# --------------------------------------------------------------------------- #
def build_nutrition(ws, plan: Plan, wr1: int, wr2: int) -> None:
    t = translator(plan.cfg.language)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1"); h1(ws["A1"], t("Nutrition — targets from current weight + principles"))
    ws.row_dimensions[1].height = 24
    widths(ws, {"A": 30, "B": 22, "C": 52})
    # current weight = last Week entry, fallback start bodyweight
    unit = t(plan.cfg.profile.units.value)
    kgf = 1.0 if plan.cfg.profile.units.value == "kg" else 0.45359237   # entries are in display units; targets are g per KG
    pos = f'SUMPRODUCT(MAX((Week!$E${wr1}:$E${wr2}<>"")*ROW(Week!$E${wr1}:$E${wr2})))-{wr1-1}'
    cw = f'IFERROR(INDEX(Week!$E${wr1}:$E${wr2},{pos}),{plan.cfg.profile.bodyweight})'
    h2(ws, 3, 1, 3, t("Targets (computed from your latest weight)"))
    nut = [
        (t("Weight used for calc, {unit}", unit=unit), f"={cw}", "0.0"),
        (t("Protein, g/day — minimum"), f"=1.8*({cw})*{kgf}", "0"),
        (t("Protein, g/day — target"), f"=2.2*({cw})*{kgf}", "0"),
        (t("Protein per meal (x4-5), g"), f"=0.3*({cw})*{kgf}", "0"),
        (t("Carbs on training days, g (~3-4 g/kg)"), f"=3.5*({cw})*{kgf}", "0"),
        (t("Fat minimum, g (~0.8 g/kg)"), f"=0.8*({cw})*{kgf}", "0"),
    ]
    r = 4
    for name, f, fmt in nut:
        td(ws.cell(r, 1), name, bold=True)
        ws.cell(r, 2, f); td(ws.cell(r, 2), center=True); ws.cell(r, 2).number_format = fmt
        ws.cell(r, 2).font = Font(FONT, 11, bold=True); fill(ws.cell(r, 2), LBLUE)
        ws.row_dimensions[r].height = 24; r += 1
    r += 1
    h2(ws, r, 1, 3, t("Principles")); r += 1
    pr = [
        (t("Rate of loss"), t("0.4-0.55 kg/wk ≈ 0.9-1.2 lb/wk (0.5-0.7% BW). Slow = keeps strength.")),
        (t("Deficit"), t("~300-500 kcal/day. No crash dieting.")),
        (t("Protein"), t("The key nutrient against strength loss in a deficit — keep near the upper target.")),
        (t("Carbs"), t("More around quality sessions (fingers, limit), less on Zone-2/rest days.")),
        (t("Collagen + vitamin C"), t("15 g gelatin/collagen + vit C, 30-60 min before finger loading.")),
        (t("Exit the deficit"), t("Return to maintenance 4-6 weeks before the goal; carb-load at the very end.")),
        (t("Supplements"), t("Creatine 3-5 g/day, caffeine 3-6 mg/kg pre-comp, vitamin D.")),
        (t("Important"), t("This is a framework. Exact calories/macros — see a sports dietitian.")),
    ]
    for k, v in pr:
        td(ws.cell(r, 1), k, bold=True); ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3); td(ws.cell(r, 2), v)
        ws.row_dimensions[r].height = 30; r += 1


# --------------------------------------------------------------------------- #
# Recovery (always)
# --------------------------------------------------------------------------- #
def build_recovery(ws, plan: Plan) -> None:
    t = translator(plan.cfg.language)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1"); h1(ws["A1"], t("Recovery & health traffic light"))
    ws.row_dimensions[1].height = 24
    widths(ws, {"A": 22, "B": 58, "C": 42})
    for i, htext in enumerate([t("Zone"), t("What it means"), t("What to do")], 1):
        th(ws.cell(3, i, htext))
    rows = [
        (GREEN, t("🟢 Green"), t("HRV at/above baseline; resting HR steady; ACWR 0.8-1.3; strength holding/rising; sleep good; fingers pain-free; losing 0.4-0.55 kg/wk."), t("Carry on as planned.")),
        (YELLOW, t("🟡 Amber"), t("HRV a few days below baseline; ACWR 1.3-1.5; fatigue 8+/10; weight stalled or dropping fast; mild finger fatigue."), t("Easy/technique day instead of limit. More food and sleep. Check rate and volume.")),
        (RED, t("🔴 Red"), t("HRV chronically low + RHR rising; ACWR >1.5; strength down 2 sessions; finger pain ≥2/3; sleep/mood/libido down; losing >0.7 kg/wk."), t("Cut volume, exit the deficit. Finger pain — pause. Persistent symptoms — see a doctor.")),
    ]
    r = 4
    for color, z, mean, act in rows:
        td(ws.cell(r, 1), z, bold=True); fill(ws.cell(r, 1), color)
        td(ws.cell(r, 2), mean); fill(ws.cell(r, 2), color)
        td(ws.cell(r, 3), act); fill(ws.cell(r, 3), color)
        ws.row_dimensions[r].height = 80; r += 1
    r += 1
    h2(ws, r, 1, 3, t("Recovery protocols")); r += 1
    rec = [
        (t("Sleep"), t("7-9 h. The main lever for CNS recovery and keeping strength in a deficit.")),
        (t("Deloads"), t("Every 3-4 weeks: volume -40-50%, intensity held (see Cycle).")),
        (t("Fingers/pulleys"), t("Always warm up progressively. Half-crimp and open grip first. Collagen slower in a deficit — add load carefully.")),
        (t("ACWR"), t("Don't ramp weekly load in jumps. >50% over the 4-week average = risk.")),
        (t("HRV / Body Battery"), t("Low in the morning → easy/technique day instead of limit or max hangs.")),
        (t("Cardio as recovery"), t("Easy Zone 2 aids recovery and burns fat with little interference.")),
    ]
    for k, v in rec:
        td(ws.cell(r, 1), k, bold=True); ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3); td(ws.cell(r, 2), v)
        ws.row_dimensions[r].height = 32; r += 1


# --------------------------------------------------------------------------- #
# Glossary (always)
# --------------------------------------------------------------------------- #
def build_glossary(ws, plan: Plan) -> None:
    t = translator(plan.cfg.language)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 28, "B": 92})
    ws.merge_cells("A1:B1"); h1(ws["A1"], t("Glossary — plain language"))
    ws.row_dimensions[1].height = 24
    for i, htext in enumerate([t("Term"), t("What it is")], 1):
        th(ws.cell(3, i, htext))
    g = [
        ("RPE", t("Subjective session hardness 1-10 (10 = max). You enter it after training.")),
        (t("sRPE load"), t("Duration × RPE. A simple measure of what a session 'cost'.")),
        (t("Weekly load"), t("Sum of session sRPE for the week. Total stress.")),
        (t("Load ramp (acute:chronic)"), t("This week's load ÷ the average of the 4 weeks before it (the current week is excluded, so the number isn't diluted by itself). Around 1.0 = steady; well above = you ramped up fast. Treat the bands as rough orientation, NOT injury prediction — see the note at the bottom.")),
        (t("Monotony"), t("Foster's measure of day-to-day sameness: mean daily load ÷ its standard deviation across the week. High monotony (>2) means every day looks alike — no hard/easy contrast. It is the combination of high load AND high monotony that tracks with overreaching, not either alone.")),
        (t("Strain"), t("Weekly load × monotony (Foster). Catches the case a load number alone misses: a big week done as seven identical days is more taxing than the same total with real rest days.")),
        (t("EWMA load"), t("Exponentially weighted moving average of weekly load — recent weeks count more, older ones decay smoothly. A less jumpy view of your chronic load than a flat 4-week mean.")),
        (t("Δ load %"), t("Plain week-over-week change in load. No model, no thresholds — just how much more or less you did than last week.")),
        (t("Relative strength (%BW)"), t("Strength relative to bodyweight. Losing weight raises it without new training.")),
        (t("Finger norm (V-target)"), t("Population guide (Lattice): the %BW max hang on a 20 mm edge that tends to match a grade. Wide spread.")),
        (t("Max hangs"), t("~7-10 s hang on an edge with added load, hard but clean (Eva Lopez method). Base finger strength.")),
        (t("Active pulls"), t("Pulling hard into a fixed edge (overcoming isometric). Safer than heavy hangs, transfers to the wall.")),
        (t("Limit bouldering"), t("Very hard boulders at your ceiling, 3-5 moves, long rests. Strength and power.")),
        (t("Power-endurance"), t("Holding high output 1-5 min under pump. For comp format; introduced late, fades fast.")),
        (t("4x4"), t("4 boulders back-to-back = a round; rest 4 min; 4 rounds.")),
        (t("Deload"), t("A lighter week every 3-4 weeks for recovery.")),
        (t("Taper"), t("Cutting volume before the goal so you arrive fresh.")),
        (t("Zone 2"), t("Easy cardio where you can still talk.")),
        (t("HRV"), t("Heart-rate variability (wearable, overnight). Falling = fatigue.")),
        (t("RED-S"), t("Energy deficiency in sport from prolonged under-fuelling at high load. Hits hormones, sleep, bone, immunity.")),
        (t("Pain 0-3"), t("An in-session signal, not a diagnosis: 0 none, 1 mild, 2 noticeable (stop signal), 3 sharp/'pop' (stop immediately). A 2+ is a prompt to consider an Injuries entry, not an automatic one.")),
        (t("Injury vs bump"), t("Log an *injury* when a structure (finger/tendon/joint) hurts — especially with no clear cause, under load, or lasting past the next session. A *bump* (a knock, skin/flapper, one-off soreness with an obvious cause) is not an injury; a Journal note is enough.")),
        (t("A2 pulley"), t("A finger pulley near the bone; the most common climber injury.")),
        (t("Half-crimp / open"), t("Grip positions; open-hand is gentler on the pulleys.")),
    ]
    r = 4
    for i, (k, v) in enumerate(g):
        td(ws.cell(r, 1), k, bold=True); td(ws.cell(r, 2), v)
        if i % 2 == 0:
            fill(ws.cell(r, 1), GREY); fill(ws.cell(r, 2), GREY)
        ws.row_dimensions[r].height = 30; r += 1
    ws.freeze_panes = "A4"

    # ---- where these numbers come from, and what they can't do ----
    r += 1
    h2(ws, r, 1, 2, t("How much to trust these numbers")); ws.row_dimensions[r].height = 20
    r += 1
    for k, v in [
        (t("What is well established"),
         t("Session-RPE (minutes × RPE) as a measure of internal training load is validated and widely used (Foster et al. 2001). "
           "Monotony and strain come from the same body of work. Finger strength as %BW correlates with climbing grade, though it "
           "explains only about half the variance — which is why it is a target band here and never a verdict.")),
        (t("What is contested"),
         t("The acute:chronic ratio has been criticised heavily since 2019: the current week is usually counted inside its own "
           "average (mathematical coupling, which manufactures correlation), the 0.8–1.3 'sweet spot' bands are largely arbitrary, "
           "and the injury-prediction figure behind them was formally challenged in the literature. This workbook excludes the "
           "current week from the chronic average to avoid the coupling, but the honest position is that the ramp number shows you "
           "WHAT CHANGED — it does not predict injury.")),
        (t("How to use them"),
         t("Treat every number here as a prompt to look, not an instruction to obey. A ramp of 1.8 with fresh fingers and good sleep "
           "may be fine; a ramp of 1.1 with sore fingers is not. Pain and how you actually feel outrank every metric on this sheet.")),
    ]:
        td(ws.cell(r, 1), k, bold=True); td(ws.cell(r, 2), v)
        ws.row_dimensions[r].height = 62; r += 1


# --------------------------------------------------------------------------- #
# How to use (always; first thing a new user should read)
# --------------------------------------------------------------------------- #
def build_howto(ws, plan: Plan) -> None:
    t = translator(plan.cfg.language)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 26, "C": 86})
    ws.merge_cells("B2:C2"); h1(ws["B2"], t("How to use — this file is your plan, tracker and advisor"))
    ws.row_dimensions[2].height = 26
    unit = t(plan.cfg.profile.units.value)
    rows = [
        (t("Two actions, that's all"),
         t("1) After every session add ONE row in Journal. 2) Once a week fill 5-6 numbers in Week "
           "(weight + recovery). Everything else — load, ACWR, finger strength, weight pace, status — computes itself.")),
        (t("Units"),
         t("All weights in this workbook are in {unit} — enter bodyweight and added load in {unit}. "
           "Percent-of-bodyweight numbers are unit-free.", unit=unit)),
        (t("Journal (after a session)"),
         t("Date (as a real date), type from the dropdown, minutes, RPE 1-10. When relevant: hang added load, "
           "best grade, pain 0-3, a note. Week number and load compute. Don't edit the grey start-date cell (B2).")),
        (t("Week (weekly check-in)"),
         t("Weight, sleep hours, fatigue, sleep quality, stress — plus HRV/resting HR if you track them. "
           "Auto columns: pace vs the planned curve, fingers %BW, gap to your grade norm, load, ACWR, "
           "completion vs plan, and a traffic-light status.")),
        (t("Dashboard"),
         t("Your landing page: headline numbers from the latest week plus an advisor that tells you what to do "
           "(adjust the deficit, back off load, swap a limit day, rest a finger).")),
        (t("Cycle & Schedules"),
         t("Cycle is the macro plan (phases, dates, deloads, planned weight). The Schedule sheets expand each "
           "phase day by day. Rearranging days is fine — keep quality days spaced and fingers ≥48h apart.")),
        (t("Injuries"),
         t("Log what affects training: structural pain (finger/tendon/joint), pain under load, anything lasting "
           "past the next session. Bumps, skin and one-off soreness with an obvious cause don't belong here.")),
        (t("How the advisor thinks"),
         t("sRPE load = minutes × RPE → weekly sum → ACWR (this week ÷ 4-week average; 0.8-1.3 ok, >1.5 risk). "
           "Plus weight pace vs plan, finger strength vs the population norm for your target grade, and pain. "
           "Finger pain overrides everything.")),
        (t("Honesty"),
         t("Norms are population guides with wide spread; this tool is not medical or coaching advice "
           "(see DISCLAIMER in the repo). When in doubt — less load, more sleep.")),
    ]
    r = 4
    for k, v in rows:
        td(ws.cell(r, 2), k, bold=True)
        td(ws.cell(r, 3), v)
        ws.row_dimensions[r].height = 46
        r += 1
