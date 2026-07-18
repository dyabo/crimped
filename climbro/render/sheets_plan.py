"""
climbro.render.sheets_plan — sheets generated directly from the engine Plan:
  - Setup   (echoes the config so the workbook is self-describing + drives some cells)
  - Cycle   (the macrocycle: phases, auto-dated, planned weight, deload flags)

These are the dynamic, per-user sheets (no hardcoded plan).
"""

from __future__ import annotations
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from .style import h1, h2, th, td, note, inp, fill, widths, GREEN, ORANGE, GREY
from ..schema import Config, format_grade, from_kg
from ..i18n import translator
from ..engine import Plan


# --------------------------------------------------------------------------- #
# Setup — human-readable echo of the configuration
# --------------------------------------------------------------------------- #
def build_setup(ws, plan: Plan) -> None:
    cfg: Config = plan.cfg
    t = translator(cfg.language)
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 32, "B": 18, "C": 3, "D": 60})
    ws.merge_cells("A1:D1"); h1(ws["A1"], t("Setup — your inputs for this plan"))
    ws.row_dimensions[1].height = 24

    gs = cfg.climbing.grade_scale
    unit = t(cfg.profile.units.value)

    def yn(b):
        return t("yes") if b else t("no")

    rows = [
        (t("Name"), cfg.profile.name or "—", ""),
        (t("Sex"), t(cfg.profile.sex.value), t("used for finger-strength norms")),
        (t("Age"), cfg.profile.age if cfg.profile.age is not None else "—", ""),
        (t("Units"), unit, ""),
        (t("Start bodyweight"), f"{cfg.profile.bodyweight} {unit}", ""),
        (t("Current grade"), format_grade(gs, cfg.climbing.current_grade), t("scale: {s}", s=gs.value)),
        (t("Target grade"), format_grade(gs, cfg.climbing.target_grade), ""),
        (t("Years climbing"), cfg.climbing.years_climbing, t("influences technique vs strength emphasis")),
        (t("Goal"), t(cfg.goal.goal.value), ""),
        (t("Start date"), cfg.goal.start_date.isoformat(), ""),
        (t("Goal date"), cfg.goal.goal_date.isoformat(), t("{weeks} weeks", weeks=plan.macro.total_weeks)),
        (t("Cut enabled"), yn(cfg.weight.enabled), ""),
        (t("Target bodyweight"), f"{cfg.weight.target_bodyweight} {unit}" if cfg.weight.enabled else "—", ""),
        (t("Training days/week"), cfg.availability.days_per_week, ""),
        (t("Fingerboard"), yn(cfg.equipment.has_fingerboard), ""),
        (t("System board"), yn(cfg.equipment.has_system_board), t("MoonBoard / Kilter / Tension")),
        (t("Gym access"), yn(cfg.equipment.has_gym), ""),
        (t("Wearable"), t(cfg.equipment.wearable.value), ""),
        (t("Mobility block"), yn(cfg.options.include_mobility), ""),
        (t("Lead sessions"), yn(cfg.options.include_lead), t("1x/week: replaces volume (base) / PE slot (peak)")),
        (t("Nutrition sheet"), yn(cfg.options.include_nutrition), ""),
    ]
    r = 3
    for label, value, desc in rows:
        td(ws.cell(r, 1), label, bold=True)
        td(ws.cell(r, 2), value, center=True)
        note(ws.cell(r, 4), desc)
        ws.row_dimensions[r].height = 22
        r += 1

    # finger-strength target headline
    r += 1
    h2(ws, r, 1, 4, t("Finger-strength target")); ws.row_dimensions[r].height = 20
    r += 1
    tgt = format_grade(gs, cfg.climbing.target_grade)
    added_disp = round((plan.norm_target_pct - 1.0) * cfg.profile.bodyweight, 1)
    msg = t("To match the {tgt} norm: ~{pct}% bodyweight (7 s, 20 mm, two hands) "
            "= about +{kg} {unit} added at {bw} {unit}. Population guide, not a hard rule.",
            tgt=tgt, pct=f"{plan.norm_target_pct*100:.0f}", kg=added_disp,
            unit=unit, bw=cfg.profile.bodyweight)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    note(ws.cell(r, 1), msg); ws.row_dimensions[r].height = 34


# --------------------------------------------------------------------------- #
# Cycle — the macrocycle table, auto-dated, with planned weight + deloads
# --------------------------------------------------------------------------- #
def build_cycle(ws, plan: Plan) -> None:
    cfg = plan.cfg
    t = translator(cfg.language)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1"); h1(ws["A1"], t("Cycle — your macrocycle (dates & planned weight computed)"))
    ws.row_dimensions[1].height = 24

    headers = [t("Wk"), t("Dates"), t("Phase"), t("Deload"), t("Focus"), t("Plan wt"), t("Sessions")]
    widths(ws, {"A": 6, "B": 20, "C": 26, "D": 9, "E": 50, "F": 10, "G": 10})
    for i, htext in enumerate(headers, 1):
        th(ws.cell(3, i, htext))
    ws.row_dimensions[3].height = 30

    start = cfg.goal.start_date
    units = t(cfg.profile.units.value)
    r = 4
    for w in plan.weeks:
        wk_start = start.toordinal() + (w.week - 1) * 7
        from datetime import date
        d0 = date.fromordinal(wk_start)
        d1 = date.fromordinal(wk_start + 6)
        td(ws.cell(r, 1), w.week, center=True)
        td(ws.cell(r, 2), f"{d0.strftime('%d.%m')}–{d1.strftime('%d.%m')}", center=True)
        td(ws.cell(r, 3), w.phase_name, bold=True)
        td(ws.cell(r, 4), t("DELOAD") if w.deload else "", center=True)
        td(ws.cell(r, 5), w.focus)
        td(ws.cell(r, 6), f"{from_kg(w.planned_weight, cfg.profile.units)} {units}" if w.planned_weight is not None else "—", center=True)
        td(ws.cell(r, 7), w.planned_session_count, center=True)
        if w.deload:
            fill(ws.cell(r, 4), ORANGE)
        if w.in_deficit:
            fill(ws.cell(r, 6), GREEN)
        ws.row_dimensions[r].height = 26
        r += 1

    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=7)
    note(ws.cell(r + 1, 1),
         t("Planned weight follows your cut curve (green = deficit weeks). Day-by-day sessions "
           "are on the phase schedule sheets. What you actually do goes in Journal / Week."))
    ws.row_dimensions[r + 1].height = 30
