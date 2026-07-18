"""
climbro.render.sheets_schedule — phase schedule sheets (day-by-day).

Expands the engine's per-week session allocation into readable schedules,
one sheet per phase group, in the format: day | session | what.
This replaces the hand-written "Расписание" sheets with generated ones.
"""

from __future__ import annotations
from .style import h1, th, td, note, fill, widths, GREEN, ORANGE, LBLUE, GREY, BLUE, FONT
from openpyxl.styles import Font, PatternFill, Alignment
from ..engine import Plan, K_REST
from ..schema import WEEKDAYS


# group consecutive phases into sheets so we don't make 6 tiny tabs
_PHASE_GROUPS = [
    ("Schedule · Assess & Base", {"assess", "base"}),
    ("Schedule · Contact", {"contact"}),
    ("Schedule · Bridge & Peak", {"bridge", "peak"}),
    ("Schedule · Taper", {"taper"}),
]


def schedule_sheets(plan: Plan) -> list[tuple[str, list]]:
    """
    Returns [(sheet_title, [WeekPlan,...]), ...] for the renderer to create.
    Only includes groups that actually have weeks (short cycles may drop phases).
    """
    out = []
    for title, keys in _PHASE_GROUPS:
        wks = [w for w in plan.weeks if w.phase_key in keys]
        if wks:
            out.append((title, wks))
    return out


def build_schedule(ws, weeks: list) -> None:
    ws.sheet_view.showGridLines = False
    title = "Schedule"
    ws.merge_cells("A1:E1"); h1(ws["A1"], _title_for(weeks))
    ws.row_dimensions[1].height = 24
    widths(ws, {"A": 4, "B": 9, "C": 7, "D": 22, "E": 80})

    r = 3
    for w in weeks:
        # week header band
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        band = ws.cell(r, 1, f"Week {w.week} · {w.phase_name}" + ("  ·  DELOAD" if w.deload else "")
                       + (f"  ·  planned {w.planned_weight}kg" if w.planned_weight is not None else ""))
        band.font = Font(FONT, 11, bold=True, color="9C5700" if w.deload else "FFFFFF")
        band.fill = PatternFill("solid", fgColor=ORANGE if w.deload else BLUE)
        band.alignment = Alignment("left", "center")
        ws.row_dimensions[r].height = 20
        r += 1
        # column header
        for i, htext in enumerate(["✓", "Day", "Type", "Session", "What"], 1):
            th(ws.cell(r, i, htext))
        ws.row_dimensions[r].height = 16
        r += 1
        # one row per training day (rest days included as grey)
        train = {s.day: s for s in w.sessions}
        for d in WEEKDAYS:
            s = train.get(d)
            if s is None:
                continue  # not a training weekday for this user
            is_rest = s.kind == K_REST
            td(ws.cell(r, 1), "✓" if is_rest else "☐", center=True)
            td(ws.cell(r, 2), d, bold=True)
            td(ws.cell(r, 3), "rest" if is_rest else ("quality" if s.quality else "support"), center=True)
            td(ws.cell(r, 4), s.kind, bold=True)
            td(ws.cell(r, 5), s.intent)
            if is_rest:
                for c in range(1, 6):
                    fill(ws.cell(r, c), GREY)
            elif s.quality:
                fill(ws.cell(r, 4), GREEN)
            else:
                fill(ws.cell(r, 4), LBLUE)
            ws.row_dimensions[r].height = 28
            r += 1
        # focus line
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        note(ws.cell(r, 1), "Focus: " + w.focus)
        ws.row_dimensions[r].height = 24
        r += 2

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    note(ws.cell(r, 1),
         "Quality days (green) = fingers/limit, kept fresh and ≥48h apart. Support = "
         "technique/strength/cardio. Rearranging days won't break anything — keep the spacing.")
    ws.row_dimensions[r].height = 32


def _title_for(weeks: list) -> str:
    keys = {w.phase_key for w in weeks}
    for title, group in _PHASE_GROUPS:
        if keys & group:
            return title
    return "Schedule"
