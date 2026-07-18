"""
climbro.render.style — shared styling helpers for all sheets.

Centralizes the palette, fonts and cell helpers so every sheet looks consistent
and the sheet modules stay focused on content/formulas.
"""

from __future__ import annotations
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"

# palette
NAVY = "1F3A5F"
BLUE = "2E6DA4"
LBLUE = "D6E4F0"
GREEN = "C8E6C9"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
RED = "F8CBAD"
GREY = "F2F2F2"
TEAL = "D7EEF0"

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def solid(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def h1(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(FONT, 15, bold=True, color="FFFFFF")
    cell.fill = solid(NAVY)
    cell.alignment = Alignment("left", "center")


def h2(ws, row: int, c1: int, c2: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row, c1, text)
    cell.font = Font(FONT, 12, bold=True, color="FFFFFF")
    cell.fill = solid(BLUE)
    cell.alignment = Alignment("left", "center")


def th(cell, text=None) -> None:
    if text is not None:
        cell.value = text
    cell.font = Font(FONT, 9, bold=True, color="FFFFFF")
    cell.fill = solid(BLUE)
    cell.alignment = Alignment("center", "center", wrap_text=True)
    cell.border = BORDER


def td(cell, text=None, center=False, bold=False, size=10) -> None:
    if text is not None:
        cell.value = text
    cell.font = Font(FONT, size, bold=bold)
    cell.border = BORDER
    cell.alignment = Alignment("center" if center else "left", "top", wrap_text=True)


def fill(cell, color: str) -> None:
    cell.fill = solid(color)


def inp(cell) -> None:
    """Mark a cell as user-input (yellow)."""
    cell.fill = solid(YELLOW)
    cell.border = BORDER
    cell.alignment = Alignment("center", "center")


def note(cell, text=None) -> None:
    if text is not None:
        cell.value = text
    cell.font = Font(FONT, 10, italic=True, color="555555")
    cell.alignment = Alignment("left", "top", wrap_text=True)


def widths(ws, mapping: dict) -> None:
    """mapping: {'A': 12, 'B': 30, ...}"""
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w
