"""
climbro.render.sheets_tracking — the analytics sheets (formula-driven):
  - Journal   (per-session log; sRPE load computed)
  - Week      (weekly check-in; pace, fingers %BW, ACWR, plan-vs-actual, status)
  - Injuries  (log + auto day-count + active flag)

Parameterized by the Plan: number of weeks, target finger-strength norm,
planned-weight curve (written as constants), and which recovery metrics exist.

Formula idioms (proven cross-engine / Google Sheets safe):
  - last value aligned to last weight check-in:  INDEX(col, SUMPRODUCT(MAX((W<>"")*ROW(W)))-off)
  - group max in a week:                          SUMPRODUCT(MAX((week=W)*values))
  - weekly load:                                  SUMIFS ; ACWR = week / 4-week avg
"""

from __future__ import annotations
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill
from .style import (h1, h2, th, td, note, inp, fill, widths,
                    FONT, RED, GREEN, YELLOW, ORANGE, GREY, LBLUE)
from openpyxl.styles import Font, Alignment
from ..engine import Plan
from ..schema import from_kg, Units
from ..i18n import translator

# session-type dropdown (English keys; translated per-language at render time)
SESSION_TYPES = ["Max hangs", "Active pulls", "Limit bouldering", "Volume / technique",
                 "Strength on the wall", "Power-endurance", "Lead climbing", "Gym strength",
                 "Zone 2 cardio", "Rest / mobility"]


# --------------------------------------------------------------------------- #
# Journal
# --------------------------------------------------------------------------- #
def build_journal(ws, plan: Plan, jr1: int, jr2: int) -> None:
    t = translator(plan.cfg.language)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:K1"); h1(ws["A1"], t("Journal — one row per session"))
    ws.row_dimensions[1].height = 24
    unit = t(plan.cfg.profile.units.value)
    headers = [t("Date"), t("Wk"), t("Type"), t("Min"), t("RPE 1-10"), t("Load"),
               t("Hang +{unit}", unit=unit), t("Grade V"), t("Volume"), t("Pain 0-3"), t("Notes")]
    w = {"A": 11, "B": 6, "C": 20, "D": 8, "E": 9, "F": 9, "G": 10, "H": 8, "I": 12, "J": 9, "K": 34}
    widths(ws, {get_column_letter(i + 1): list(w.values())[i] for i in range(len(w))})
    for i, htext in enumerate(headers, 1):
        th(ws.cell(4, i, htext))
    ws.row_dimensions[4].height = 38

    inputs = {1, 3, 4, 5, 7, 8, 9, 10, 11}
    for r in range(jr1, jr2 + 1):
        for c in range(1, 12):
            td(ws.cell(r, c))
            if c in inputs:
                fill(ws.cell(r, c), YELLOW)
        ws.cell(r, 1).number_format = "dd.mm.yyyy"
        # week number from date vs Setup start date is awkward across sheets;
        # use the plan start date constant instead (written into B2 helper).
        ws.cell(r, 2, f'=IF(A{r}="","",INT((A{r}-$B$2)/7)+1)')
        ws.cell(r, 2).alignment = Alignment("center", "center")
        ws.cell(r, 6, f'=IF(OR(D{r}="",E{r}=""),"",D{r}*E{r})')
        ws.cell(r, 6).alignment = Alignment("center", "center")
    # hidden helper: plan start date in B2 (used by week-number formula)
    ws["A2"] = t("start date (don't edit) →"); note(ws["A2"])
    ws["B2"] = plan.cfg.goal.start_date
    ws["B2"].number_format = "dd.mm.yyyy"
    fill(ws["B2"], GREY)
    ws.freeze_panes = "A5"

    session_types = [t(x) for x in SESSION_TYPES]
    dv = DataValidation(type="list", formula1='"' + ",".join(session_types) + '"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"C{jr1}:C{jr2}")
    dvr = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
    ws.add_data_validation(dvr); dvr.add(f"E{jr1}:E{jr2}")
    dvp = DataValidation(type="whole", operator="between", formula1="0", formula2="3", allow_blank=True)
    ws.add_data_validation(dvp); dvp.add(f"J{jr1}:J{jr2}")
    ws.conditional_formatting.add(f"J{jr1}:J{jr2}",
        CellIsRule(operator="greaterThanOrEqual", formula=["2"], fill=PatternFill("solid", fgColor=RED)))


# --------------------------------------------------------------------------- #
# Week
# --------------------------------------------------------------------------- #
# fixed column map (kept stable regardless of metric flags; unused metric cols greyed)
#  1A Wk 2B Dates 3C Phase 4D Deload 5E Weight 6F PlanWt 7G dPlan 8H dWk 9I Pace
# 10J HRV 11K HRV-base 12L RHR 13M RHR-base 14N Sleep h 15O Fatigue
# 16P Sessions 17Q Load 18R Chronic 19S ACWR 20T Fingers%BW 21U ToTarget
# 22V GradeV 23W MaxPain 24X Status 25Y SleepQ 26Z Stress 27AA PlanSes 28AB Done%
def build_week(ws, plan: Plan, jr1: int, jr2: int) -> None:
    t = translator(plan.cfg.language)
    ws.sheet_view.showGridLines = False
    n = len(plan.weeks)
    WR1 = 5
    WR2 = WR1 + n - 1
    JB = f"Journal!$B${jr1}:$B${jr2}"
    JD = f"Journal!$D${jr1}:$D${jr2}"
    JF = f"Journal!$F${jr1}:$F${jr2}"
    JG = f"Journal!$G${jr1}:$G${jr2}"
    JH = f"Journal!$H${jr1}:$H${jr2}"
    JJ = f"Journal!$J${jr1}:$J${jr2}"
    norm = plan.norm_target_pct           # literal target finger-strength norm
    m = plan.metrics
    units = plan.cfg.profile.units
    # weekly-rate thresholds in the user's display units (internally defined in kg)
    fast_thr = 0.7 if units == Units.KG else 1.5     # 'cutting too fast'
    gain_thr = 0.1 if units == Units.KG else 0.2     # 'gaining'
    p_fast, p_gain, p_ok = t("fast"), t("gain"), t("ok")  # pace labels (also used in CF below)

    ws.merge_cells("A1:AB1"); h1(ws["A1"], t("Week — fill yellow; the rest computes"))
    ws.row_dimensions[1].height = 24

    # baseline block (only meaningful if a wearable tracks HRV/RHR)
    ws["A2"] = t("Baseline HRV:"); ws["A2"].font = Font(FONT, 10, bold=True)
    inp(ws["B2"])
    ws["C2"] = t("Baseline resting HR:"); ws["C2"].font = Font(FONT, 10, bold=True)
    inp(ws["D2"])
    if not (m["hrv"] or m["resting_hr"]):
        note(ws["F2"], t("No wearable selected — HRV / resting-HR columns are optional."))
        ws.merge_cells("F2:AB2")

    headers = [t("Wk"), t("Dates"), t("Phase"), t("Deload"), t("Weight"), t("Plan wt"), t("Δ plan"), t("Δ/wk"), t("Pace"),
               t("HRV"), t("HRV−base"), t("Rest HR"), t("RHR−base"), t("Sleep h"), t("Fatigue 1-10"),
               t("Sessions"), t("Load"), t("Chronic"), t("ACWR"), t("Fingers %BW"), t("To target"),
               t("Grade V"), t("Max pain"), t("Week status"), t("Sleep Q 1-5"), t("Stress 1-10"), t("Plan ses"), t("Done %")]
    cw = [6, 13, 22, 8, 8, 9, 9, 8, 9, 7, 9, 9, 9, 7, 9, 7, 9, 9, 7, 10, 9, 8, 9, 24, 9, 9, 9, 8]
    widths(ws, {get_column_letter(i + 1): cw[i] for i in range(len(cw))})
    for i, htext in enumerate(headers, 1):
        th(ws.cell(4, i, htext))
    ws.row_dimensions[4].height = 46

    metric_input_cols = set()
    if m["hrv"]:
        metric_input_cols.add(10)
    if m["resting_hr"]:
        metric_input_cols.add(12)
    base_input = {5, 14, 15, 25, 26}  # weight, sleep h, fatigue, sleep Q, stress
    input_cols = base_input | metric_input_cols

    for i, wp in enumerate(plan.weeks):
        r = WR1 + i
        td(ws.cell(r, 1), wp.week, center=True)
        # dates as constants from the plan
        from datetime import date
        d0 = date.fromordinal(plan.cfg.goal.start_date.toordinal() + (wp.week - 1) * 7)
        td(ws.cell(r, 2), d0.strftime("%d.%m"), center=True)
        td(ws.cell(r, 3), wp.phase_name.split(" — ")[0].split(" · ")[0], center=True)
        td(ws.cell(r, 4), t("DELOAD") if wp.deload else "", center=True)
        if wp.deload:
            fill(ws.cell(r, 4), ORANGE)
        # inputs
        for c in input_cols:
            inp(ws.cell(r, c))
        # grey out metric cols that aren't tracked
        if not m["hrv"]:
            for c in (10, 11):
                fill(ws.cell(r, c), GREY); td(ws.cell(r, c))
        if not m["resting_hr"]:
            for c in (12, 13):
                fill(ws.cell(r, c), GREY); td(ws.cell(r, c))
        # plan weight as a constant (None -> blank, no-cut)
        if wp.planned_weight is not None:
            td(ws.cell(r, 6), from_kg(wp.planned_weight, units), center=True); ws.cell(r, 6).number_format = "0.0"
        else:
            td(ws.cell(r, 6), "", center=True)
        # Δ to plan / Δ per week / pace
        td(ws.cell(r, 7)); ws.cell(r, 7, f'=IF(OR(E{r}="",F{r}=""),"",E{r}-F{r})'); ws.cell(r, 7).number_format = "0.0"; ws.cell(r, 7).alignment = Alignment("center", "center")
        if r > WR1:
            ws.cell(r, 8, f'=IF(OR(E{r}="",E{r-1}=""),"",E{r}-E{r-1})')
        td(ws.cell(r, 8)); ws.cell(r, 8).number_format = "0.0"; ws.cell(r, 8).alignment = Alignment("center", "center")
        ws.cell(r, 9, f'=IF(H{r}="","",IF(H{r}<-{fast_thr},"{p_fast}",IF(H{r}>{gain_thr},"{p_gain}","{p_ok}")))'); td(ws.cell(r, 9)); ws.cell(r, 9).alignment = Alignment("center", "center")
        # HRV / RHR deviations (blank if not tracked or no baseline)
        ws.cell(r, 11, f'=IF(OR(J{r}="",$B$2=""),"",J{r}-$B$2)'); ws.cell(r, 11).alignment = Alignment("center", "center")
        ws.cell(r, 13, f'=IF(OR(L{r}="",$D$2=""),"",L{r}-$D$2)'); ws.cell(r, 13).alignment = Alignment("center", "center")
        # sessions / load / chronic / ACWR
        ws.cell(r, 16, f'=COUNTIFS({JB},A{r},{JD},">0")'); ws.cell(r, 16).alignment = Alignment("center", "center")
        ws.cell(r, 17, f'=SUMIFS({JF},{JB},A{r})'); ws.cell(r, 17).alignment = Alignment("center", "center")
        lo = max(WR1, r - 3)
        ws.cell(r, 18, f'=IF(Q{r}=0,"",AVERAGE(Q{lo}:Q{r}))'); ws.cell(r, 18).number_format = "0"; ws.cell(r, 18).alignment = Alignment("center", "center")
        ws.cell(r, 19, f'=IF(OR(R{r}="",R{r}=0),"",Q{r}/R{r})'); ws.cell(r, 19).number_format = "0.00"; ws.cell(r, 19).alignment = Alignment("center", "center")
        # fingers %BW, gap to target, grade, pain
        ws.cell(r, 20, f'=IF(OR(E{r}="",COUNTIFS({JB},A{r},{JG},">0")=0),"",(E{r}+SUMPRODUCT(MAX(({JB}=A{r})*{JG})))/E{r})'); ws.cell(r, 20).number_format = "0.0%"; ws.cell(r, 20).alignment = Alignment("center", "center")
        ws.cell(r, 21, f'=IF(T{r}="","",{norm}-T{r})'); ws.cell(r, 21).number_format = "0.0%"; ws.cell(r, 21).alignment = Alignment("center", "center")
        ws.cell(r, 22, f'=IF(COUNTIFS({JB},A{r},{JH},">0")=0,"",SUMPRODUCT(MAX(({JB}=A{r})*{JH})))'); ws.cell(r, 22).alignment = Alignment("center", "center")
        ws.cell(r, 23, f'=SUMPRODUCT(MAX(({JB}=A{r})*{JJ}))'); ws.cell(r, 23).alignment = Alignment("center", "center")
        # plan sessions (constant from engine) + done %
        td(ws.cell(r, 27), wp.planned_session_count, center=True)
        ws.cell(r, 28, f'=IF(OR(P{r}=0,AA{r}=0,AA{r}=""),"",P{r}/AA{r})'); ws.cell(r, 28).number_format = "0%"; ws.cell(r, 28).alignment = Alignment("center", "center")
        # status (priority: pain > load > weight > recovery)
        ws.cell(r, 24,
            f'=IF(W{r}>=2,"{t("🔴 Finger pain")}",'
            f'IF(AND(S{r}<>"",S{r}>1.5),"{t("🔴 Load spike")}",'
            f'IF(AND(H{r}<>"",H{r}<-{fast_thr}),"{t("🔴 Cutting too fast")}",'
            f'IF(AND(K{r}<>"",K{r}<-5,M{r}<>"",M{r}>0),"{t("🟡 Fatigue (HRV↓)")}",'
            f'IF(AND(O{r}<>"",O{r}>=8),"{t("🟡 High fatigue")}",'
            f'IF(AND(Z{r}<>"",Z{r}>=8),"{t("🟡 High stress")}",'
            f'IF(AND(Y{r}<>"",Y{r}<=2),"{t("🟡 Poor sleep")}",'
            f'IF(AND(G{r}<>"",G{r}>1),"{t("🟡 Behind on weight")}",'
            f'IF(P{r}=0,"—","{t("🟢 On track")}")))))))))')
        td(ws.cell(r, 24))
        ws.row_dimensions[r].height = 26

    ws.freeze_panes = "E5"
    last = WR2
    ws.conditional_formatting.add(f"I{WR1}:I{last}", CellIsRule(operator="equal", formula=[f'"{p_fast}"'], fill=PatternFill("solid", fgColor=RED)))
    ws.conditional_formatting.add(f"I{WR1}:I{last}", CellIsRule(operator="equal", formula=[f'"{p_ok}"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"S{WR1}:S{last}", CellIsRule(operator="greaterThan", formula=["1.5"], fill=PatternFill("solid", fgColor=RED)))
    ws.conditional_formatting.add(f"S{WR1}:S{last}", CellIsRule(operator="between", formula=["0.8", "1.3"], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"U{WR1}:U{last}", CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"W{WR1}:W{last}", CellIsRule(operator="greaterThanOrEqual", formula=["2"], fill=PatternFill("solid", fgColor=RED)))
    for emoji, color in (("🔴", RED), ("🟡", YELLOW), ("🟢", GREEN)):
        ws.conditional_formatting.add(f"X{WR1}:X{last}",
            FormulaRule(formula=[f'ISNUMBER(SEARCH("{emoji}",X{WR1}))'], fill=PatternFill("solid", fgColor=color)))
    ws.conditional_formatting.add(f"AB{WR1}:AB{last}", CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=PatternFill("solid", fgColor=GREEN)))
    ws.conditional_formatting.add(f"AB{WR1}:AB{last}", CellIsRule(operator="lessThan", formula=["0.7"], fill=PatternFill("solid", fgColor=ORANGE)))

    return WR1, WR2  # for dashboard wiring


# --------------------------------------------------------------------------- #
# Injuries
# --------------------------------------------------------------------------- #
def build_injuries(ws, lang: str = "en", tr1: int = 4, tr2: int = 33) -> str:
    t = translator(lang)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:I1"); h1(ws["A1"], t("Injuries & niggles — log + rehab"))
    ws.row_dimensions[1].height = 24
    headers = [t("Start date"), t("Area / location"), t("Type"), t("Severity 0-3"), t("Status"),
               t("Days"), t("What I'm doing (rehab)"), t("Resolved date"), t("Notes")]
    cw = [12, 26, 16, 12, 14, 8, 30, 14, 28]
    widths(ws, {get_column_letter(i + 1): cw[i] for i in range(len(cw))})
    for i, htext in enumerate(headers, 1):
        th(ws.cell(3, i, htext))
    ws.row_dimensions[3].height = 30

    inputs = {1, 2, 3, 4, 5, 7, 8, 9}
    for r in range(tr1, tr2 + 1):
        for c in range(1, 10):
            td(ws.cell(r, c), center=(c in (4, 6)))
            if c in inputs:
                fill(ws.cell(r, c), YELLOW)
        ws.cell(r, 1).number_format = "dd.mm.yyyy"
        ws.cell(r, 8).number_format = "dd.mm.yyyy"
        ws.cell(r, 6, f'=IF(A{r}="","",IF(H{r}="",TODAY()-A{r},H{r}-A{r}))'); ws.cell(r, 6).alignment = Alignment("center", "center")
    ws.freeze_panes = "A4"

    type_opts = ",".join(t(x) for x in ("Finger/tendon", "Elbow", "Shoulder", "Wrist", "Knee", "Back", "Other"))
    st_active, st_rehab, st_resolved = t("Active"), t("Rehab"), t("Resolved")
    dvt = DataValidation(type="list", formula1=f'"{type_opts}"', allow_blank=True)
    ws.add_data_validation(dvt); dvt.add(f"C{tr1}:C{tr2}")
    dvs = DataValidation(type="list", formula1=f'"{st_active},{st_rehab},{st_resolved}"', allow_blank=True)
    ws.add_data_validation(dvs); dvs.add(f"E{tr1}:E{tr2}")
    dvts = DataValidation(type="whole", operator="between", formula1="0", formula2="3", allow_blank=True)
    ws.add_data_validation(dvts); dvts.add(f"D{tr1}:D{tr2}")
    ws.conditional_formatting.add(f"E{tr1}:E{tr2}", CellIsRule(operator="equal", formula=[f'"{st_active}"'], fill=PatternFill("solid", fgColor=RED)))
    ws.conditional_formatting.add(f"E{tr1}:E{tr2}", CellIsRule(operator="equal", formula=[f'"{st_rehab}"'], fill=PatternFill("solid", fgColor=YELLOW)))
    ws.conditional_formatting.add(f"E{tr1}:E{tr2}", CellIsRule(operator="equal", formula=[f'"{st_resolved}"'], fill=PatternFill("solid", fgColor=GREEN)))
    ws.merge_cells(start_row=tr2 + 2, start_column=1, end_row=tr2 + 2, end_column=9)
    note(ws.cell(tr2 + 2, 1),
         t("Log what affects your training: a structural niggle (finger / tendon / joint), pain under load, "
           "or anything lasting past the next session — set status Active and unload that area. Do NOT log bumps, "
           "skin / flappers, or one-off soreness with an obvious cause (a Journal note is enough). When unsure: "
           "unexplained finger/tendon pain, or lost grip function → log it. 'Days' auto-counts (to resolved date "
           "or today); Active/Rehab rows feed the Dashboard advisor."))
    ws.row_dimensions[tr2 + 2].height = 52
    # injury-active count formula string for the dashboard to reuse (translated status values)
    return f'COUNTIF(Injuries!$E${tr1}:$E${tr2},"{st_active}")+COUNTIF(Injuries!$E${tr1}:$E${tr2},"{st_rehab}")'
