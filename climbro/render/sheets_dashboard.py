"""
climbro.render.sheets_dashboard — Dashboard (KPIs + advisor) and Charts.

Reads the Week sheet via "last value" idioms (Google-Sheets safe) and emits
concrete-action advice. Metric lines adapt to whether a wearable is present.
"""

from __future__ import annotations
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment
from .style import (h1, h2, th, td, note, fill, widths, FONT, GREY, RED, GREEN, ORANGE, TEAL)
from ..engine import Plan


def _lastw(col: str, wr1: int, wr2: int) -> str:
    """Value of `col` at the last week with a bodyweight entry (col E)."""
    pos = f'SUMPRODUCT(MAX((Week!$E${wr1}:$E${wr2}<>"")*ROW(Week!$E${wr1}:$E${wr2})))-{wr1-1}'
    return f'INDEX(Week!${col}${wr1}:${col}${wr2},{pos})'


def _lastne(col: str, wr1: int, wr2: int) -> str:
    """Last non-empty value of `col` itself (for columns that fill irregularly)."""
    pos = f'SUMPRODUCT(MAX((Week!${col}${wr1}:${col}${wr2}<>"")*ROW(Week!${col}${wr1}:${col}${wr2})))-{wr1-1}'
    return f'INDEX(Week!${col}${wr1}:${col}${wr2},{pos})'


def _phase_today(plan: Plan, wr1: int, wr2: int) -> str:
    """Phase by calendar date (clamped to the plan), independent of check-ins."""
    d = plan.cfg.goal.start_date
    n = len(plan.weeks)
    return (f'=INDEX(Week!$C${wr1}:$C${wr2},'
            f'MIN({n},MAX(1,INT((TODAY()-DATE({d.year},{d.month},{d.day}))/7)+1)))')


def build_dashboard(ws, plan: Plan, wr1: int, wr2: int, jr2: int = 10000) -> None:
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 30, "B": 18, "C": 3, "D": 64})
    ws.merge_cells("A1:D1"); h1(ws["A1"], "Dashboard — where you are and what to do")
    ws.row_dimensions[1].height = 26

    cfg = plan.cfg
    unit = cfg.profile.units.value
    start_bw = cfg.profile.bodyweight                 # display units
    target_bw = cfg.weight.target_bodyweight if cfg.weight.enabled else None  # display units
    rate_txt = "-0.5 kg/wk" if unit == "kg" else "-1.1 lb/wk"
    norm = plan.norm_target_pct
    tgt_v = plan.target_v
    inj = getattr(plan, "_inj_active_formula", '0')

    h2(ws, 3, 1, 2, "Now (last completed week)")
    ws.merge_cells("C3:D3"); h2(ws, 3, 3, 4, "Reading")

    kpi = [
        ("Current weight", f'=IFERROR({_lastw("E",wr1,wr2)},"—")', "0.0", "From the latest check-in"),
        ("Lost so far", f'=IFERROR({start_bw}-{_lastw("E",wr1,wr2)},"—")', "0.0", "From start"),
    ]
    if target_bw is not None:
        kpi.append(("Left to target", f'=IFERROR({_lastw("E",wr1,wr2)}-{target_bw},"—")', "0.0", "≤0 = weight goal met"))
    kpi += [
        ("Current phase", _phase_today(plan, wr1, wr2), "@", "Where you are in the plan (by date)"),
        ("Fingers %BW", f'=IFERROR({_lastne("T",wr1,wr2)},"—")', "0.0%", "(bw+hang)/bw"),
        ("To V-target norm", f'=IFERROR({_lastne("U",wr1,wr2)},"—")', "0.0%", f"≤0 = V{tgt_v} finger norm reached"),
        ("Best grade (wk)", f'=IFERROR({_lastne("V",wr1,wr2)},"—")', "0", "Max in a week"),
        ("Week load (sRPE)", f'=IFERROR({_lastw("Q",wr1,wr2)},"—")', "0", "Sum of min×RPE"),
        ("ACWR (overload risk)", f'=IFERROR({_lastne("S",wr1,wr2)},"—")', "0.00", "0.8–1.3 ok · >1.5 risky"),
        ("Week status", f'=IFERROR({_lastw("X",wr1,wr2)},"—")', "@", "Composite traffic light"),
        ("Weeks logged", f'=COUNT(Week!$E${wr1}:$E${wr2})', "0", "Check-ins so far"),
        ("Sessions logged", f'=COUNTIFS(Journal!$D$5:$D${jr2},">0")', "0", "Journal rows"),
        ("Week completion", f'=IFERROR({_lastw("AB",wr1,wr2)},"—")', "0%", "Actual ÷ planned sessions"),
        ("Active injuries", f'={inj}', "0", "Active + rehab (see Injuries)"),
    ]
    r = 4
    left_to_target_row = None
    acwr_row = None
    inj_row = None
    for name, f, fmt, desc in kpi:
        td(ws.cell(r, 1), name, bold=True)
        ws.cell(r, 2, f); td(ws.cell(r, 2), center=True); ws.cell(r, 2).number_format = fmt
        ws.cell(r, 2).font = Font(FONT, 12, bold=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        td(ws.cell(r, 3), desc)
        if name == "Left to target":
            left_to_target_row = r
        if name == "ACWR (overload risk)":
            acwr_row = r
        if name == "Active injuries":
            inj_row = r
        if (r % 2) == 0:
            for c in (1, 2, 3):
                fill(ws.cell(r, c), GREY)
        ws.row_dimensions[r].height = 26
        r += 1

    if left_to_target_row:
        ws.conditional_formatting.add(f"B{left_to_target_row}", CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor=GREEN)))
    if acwr_row:
        ws.conditional_formatting.add(f"B{acwr_row}", CellIsRule(operator="greaterThan", formula=["1.5"], fill=PatternFill("solid", fgColor=RED)))
    if inj_row:
        ws.conditional_formatting.add(f"B{inj_row}", CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=PatternFill("solid", fgColor=RED)))

    # ---- advisor (concrete actions) ----
    r += 1
    h2(ws, r, 1, 4, "Advisor"); ws.row_dimensions[r].height = 20
    r += 1
    cw = _lastw("E", wr1, wr2)
    advice = []
    if target_bw is not None:
        advice.append(
            f'=IFERROR("Weight: down "&TEXT({start_bw}-{cw},"0.0")&" of "&TEXT({start_bw}-{target_bw},"0.0")&" kg, "&TEXT({cw}-{target_bw},"0.0")&" to go. "'
            f'&IF({_lastw("G",wr1,wr2)}>1,"Behind the planned curve — add ~100-150 kcal deficit or 1 Zone-2 session.",'
            f'IF({_lastw("G",wr1,wr2)}<-1,"Faster than planned — raise calories; aim {rate_txt} to keep strength.","On the planned curve.")),"Weight: enter a bodyweight in Week.")'
        )
    advice += [
        f'=IFERROR("Fingers: "&TEXT({_lastne("T",wr1,wr2)},"0.0%")&" BW; to the V{tgt_v} norm ("&TEXT({norm},"0%")&") "'
        f'&IF({_lastne("U",wr1,wr2)}<=0,"— reached. Now convert it on the wall (technique, limit).",'
        f'"need "&TEXT({_lastne("U",wr1,wr2)}*100,"0.0")&" pts (~"&TEXT({_lastne("U",wr1,wr2)}*{cw},"0.0")&" {unit}). Keep 2 finger sessions/wk, add load slowly."),"Fingers: log a max hang in Journal.")',
        f'=IFERROR("Fatigue: ACWR "&TEXT({_lastne("S",wr1,wr2)},"0.00")&". "'
        f'&IF({_lastne("S",wr1,wr2)}>1.5,"Sharp jump — next week drop 1 power-endurance/volume session and add a rest day.",'
        f'IF({_lastne("S",wr1,wr2)}<0.8,"Low — you can add 1 volume/technique session.","In the optimal band, hold steady.")),"Fatigue: ACWR builds after ~2-3 logged weeks.")',
        f'=IF({inj}>0,"⚠ Active injuries: "&TEXT({inj},"0")&" — follow the rehab in Injuries, don'+chr(39)+'t load the area, swap the affected sessions.","Injuries: none active.")',
        f'=IFERROR(IF({_lastw("W",wr1,wr2)}>=2,"Finger pain "&TEXT({_lastw("W",wr1,wr2)},"0")&"/3 — rest fingers, switch to legs/cardio/mobility, add an Injuries row.",'
        f'IF(OR(AND({_lastw("Y",wr1,wr2)}<>"",{_lastw("Y",wr1,wr2)}<=2),AND({_lastw("Z",wr1,wr2)}<>"",{_lastw("Z",wr1,wr2)}>=8)),"Sleep/stress are down — swap the next limit day for a technique day, prioritise sleep.","Recovery: sleep and stress look fine.")),"Log pain, sleep and stress to track recovery.")',
        f'=IFERROR("Completion: "&TEXT({_lastw("AB",wr1,wr2)},"0%")&". "&IF({_lastw("AB",wr1,wr2)}<0.7,"Week under-done — don'+chr(39)+'t cram it back; add volume gradually (+10-15%).","Volume on plan."),"Completion shows once sessions + plan exist.")',
        f'=IFERROR("Overall last-week status: "&{_lastw("X",wr1,wr2)},"Status appears after your first check-in.")',
    ]
    for a in advice:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, a); td(ws.cell(r, 1)); fill(ws.cell(r, 1), TEAL)
        ws.row_dimensions[r].height = 34
        r += 1


def build_charts(ws, plan: Plan, wr1: int, wr2: int) -> None:
    import math
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:S1"); h1(ws["A1"], "Charts (populate as you log)")
    ws.row_dimensions[1].height = 24
    week = next((s for s in ws.parent.worksheets if s.title == "Week"), None)
    if week is None:
        return
    hdr_row = 4
    cfg = plan.cfg
    m = plan.metrics

    def line(title, cols, anchor, ymin=None, ymax=None, yfmt=None):
        """One line chart; each col in `cols` becomes its own named series.
        ymin/ymax pin the y-axis to the plan's real range (no more 0-based squash)."""
        ch = LineChart(); ch.title = title; ch.height = 7.5; ch.width = 14; ch.style = 2
        for c in cols:
            ref = Reference(week, min_col=c, max_col=c, min_row=hdr_row, max_row=wr2)
            ch.add_data(ref, titles_from_data=True)
        ch.set_categories(Reference(week, min_col=1, min_row=wr1, max_row=wr2))
        ch.x_axis.title = "Week"
        ch.x_axis.delete = False; ch.y_axis.delete = False
        if ymin is not None:
            ch.y_axis.scaling.min = ymin
        if ymax is not None:
            ch.y_axis.scaling.max = ymax
        if yfmt is not None:
            ch.y_axis.numFmt = yfmt
        ws.add_chart(ch, anchor)

    # ----- data-driven axis bounds (known at generation time) -----
    # Weight: tight band around the planned start→target range, not a 0-based axis.
    start = cfg.profile.bodyweight
    has_cut = cfg.weight.enabled and cfg.weight.target_bodyweight is not None
    target = cfg.weight.target_bodyweight if has_cut else start
    w_lo, w_hi = min(start, target), max(start, target)
    w_min, w_max = math.floor(w_lo - 4), math.ceil(w_hi + 4)

    # Fingers %BW: focus on [current strength .. target norm] with a little padding.
    norm = plan.norm_target_pct
    cur_pct = (cfg.bw_kg + cfg.max_hang_kg) / cfg.bw_kg if cfg.max_hang_kg is not None else None
    f_lo = min(cur_pct, norm) if cur_pct is not None else norm - 0.4
    f_min = max(0.5, math.floor((f_lo - 0.1) * 10) / 10)
    f_max = math.ceil((norm + 0.1) * 10) / 10

    # Grade: around current → target.
    g_min, g_max = max(0, plan.current_v - 2), plan.target_v + 2
    # Sessions: 0 → a bit above the busiest planned week.
    s_max = max((w.planned_session_count for w in plan.weeks), default=6) + 1

    # ----- charts (2-up: left column A, right column K) -----
    line("Weight: actual vs plan" if has_cut else "Weight",
         [5, 6] if has_cut else [5], "A3", w_min, w_max)
    line(f"Fingers %BW (target {norm*100:.0f}%)", [20], "K3", f_min, f_max, yfmt="0%")

    line("Best grade / week (V)", [22], "A18", g_min, g_max)
    line("Weekly load: acute vs chronic (sRPE)", [17, 18], "K18", 0)

    line("ACWR — overload risk", [19], "A33", 0, 2)
    line("Sessions: planned vs done", [27, 16], "K33", 0, s_max)

    line("Fatigue & stress (1-10)", [15, 26], "A48", 0, 10)
    line("Sleep (h / night)", [14], "K48", 0, 12)

    line("Max finger pain (0-3)", [23], "A63", 0, 3)
    if m["hrv"] or m["resting_hr"]:
        hrv_cols = ([10] if m["hrv"] else []) + ([12] if m["resting_hr"] else [])
        line("HRV & resting HR", hrv_cols, "K63")   # auto y-axis (device-dependent range)

    note(ws.cell(80, 1),
         "Lines appear as Week fills in. Axes are pre-scaled to your plan's range; "
         "keep ACWR in the 0.8–1.3 band and finger pain at 0.")
    ws.merge_cells("A80:S80")
