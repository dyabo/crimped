"""
crimped.render — turns an engine Plan into the .xlsx workbook.

v1 build order (this file orchestrates; sheet modules do the work):
  - sheets_plan:      Setup, Cycle                      [done]
  - sheets_schedule:  per-phase day-by-day schedules    [done]
  - sheets_tracking:  Journal, Week, Injuries, Dashboard, Charts   [next]
  - sheets_static:    Mobility, Nutrition, Recovery, Glossary (flag-gated)  [next]

Optional sheets are gated by config flags (cut / wearable / mobility / nutrition).
"""

from __future__ import annotations
import os
from openpyxl import Workbook

from ..engine import Plan
from .sheets_plan import build_setup, build_cycle
from .sheets_schedule import schedule_sheets, build_schedule
from .sheets_tracking import build_journal, build_week, build_injuries
from .sheets_dashboard import build_dashboard, build_charts
from .sheets_static import build_mobility, build_nutrition, build_recovery, build_glossary, build_howto


def build_workbook(plan: Plan) -> Workbook:
    """Assemble the whole workbook in memory (no I/O). Shared by render/render_bytes."""
    wb = Workbook()
    wb.remove(wb.active)

    n = len(plan.weeks)
    jr1 = 5
    jr2 = jr1 + max(180, n * 7) - 1
    wr1 = 5
    wr2 = wr1 + n - 1
    opt = plan.cfg.options

    # order: dashboard first (landing), then how-to/plan, then logs, then reference
    ws_dash = wb.create_sheet("Dashboard")      # filled after Week exists (needs ranges)
    build_howto(wb.create_sheet("How to use"), plan)
    build_setup(wb.create_sheet("Setup"), plan)
    build_cycle(wb.create_sheet("Cycle"), plan)
    for title, weeks in schedule_sheets(plan):
        build_schedule(wb.create_sheet(_safe(title)), weeks, title, plan.cfg.language)

    build_journal(wb.create_sheet("Journal"), plan, jr1, jr2)
    build_week(wb.create_sheet("Week"), plan, jr1, jr2)
    plan._inj_active_formula = build_injuries(wb.create_sheet("Injuries"), plan.cfg.language)

    # now that Week/Injuries exist, fill dashboard + charts
    build_dashboard(ws_dash, plan, wr1, wr2, jr2)
    build_charts(wb.create_sheet("Charts"), plan, wr1, wr2)

    if opt.include_mobility:
        build_mobility(wb.create_sheet("Mobility"), plan)
    if opt.include_nutrition:
        build_nutrition(wb.create_sheet("Nutrition"), plan, wr1, wr2)
    build_recovery(wb.create_sheet("Recovery"), plan)
    build_glossary(wb.create_sheet("Glossary"), plan)
    return wb


def render(plan: Plan, path: str) -> str:
    """Write the workbook to `path` and return the path (CLI/file use)."""
    wb = build_workbook(plan)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
    return path


def render_bytes(plan: Plan) -> bytes:
    """Return the workbook as .xlsx bytes (web/browser download, tests — no disk I/O)."""
    from io import BytesIO
    buf = BytesIO()
    build_workbook(plan).save(buf)
    return buf.getvalue()


def _safe(title: str) -> str:
    # Excel sheet titles: <=31 chars, no : \ / ? * [ ]
    bad = ':\\/?*[]'
    t = "".join(c for c in title if c not in bad)
    return t[:31]


if __name__ == "__main__":
    from ..schema import from_dict
    from ..engine import build_plan
    cfg = from_dict({
        "profile": {"name": "Bo", "sex": "male", "units": "kg", "bodyweight": 83},
        "climbing": {"grade_scale": "V", "current_grade": "V5", "target_grade": "V8",
                     "years_climbing": 1, "max_hang_added": 15, "max_pullup_added": 20},
        "goal": {"goal": "competition", "start_date": "2026-06-08", "goal_date": "2026-12-28"},
        "weight": {"enabled": True, "target_bodyweight": 78},
        "availability": {"days_per_week": 6, "available_days": ["Mon", "Tue", "Wed", "Thu", "Sat", "Sun"]},
        "equipment": {"has_fingerboard": True, "has_system_board": True, "has_gym": True, "wearable": "garmin"},
        "options": {"include_mobility": True, "include_nutrition": True},
    })
    out = render(build_plan(cfg), "/mnt/user-data/outputs/_skeleton_test.xlsx")
    print("wrote", out)
