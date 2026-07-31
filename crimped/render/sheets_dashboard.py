"""
crimped.render.sheets_dashboard — Dashboard (KPIs + advisor) and Charts.

Reads the Week sheet via "last value" idioms (Google-Sheets safe) and emits
concrete-action advice. Metric lines adapt to whether a wearable is present.
"""

from __future__ import annotations
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment
from .style import (h1, h2, th, td, note, fill, widths, FONT, GREY, RED, GREEN, ORANGE, TEAL)
from ..engine import Plan
from ..i18n import translator


def _lastw(col: str, wr1: int, wr2: int) -> str:
    """Value of `col` at the last week with a bodyweight entry (col E)."""
    pos = f'SUMPRODUCT(MAX((Week!$E${wr1}:$E${wr2}<>"")*ROW(Week!$E${wr1}:$E${wr2})))-{wr1-1}'
    return f'INDEX(Week!${col}${wr1}:${col}${wr2},{pos})'


def _lastne(col: str, wr1: int, wr2: int) -> str:
    """Last non-empty value of `col` itself (for columns that fill irregularly)."""
    pos = f'SUMPRODUCT(MAX((Week!${col}${wr1}:${col}${wr2}<>"")*ROW(Week!${col}${wr1}:${col}${wr2})))-{wr1-1}'
    return f'INDEX(Week!${col}${wr1}:${col}${wr2},{pos})'


def _checkin_row(wr1: int, wr2: int) -> str:
    """Absolute row of the last COMPLETED week = the last weekly check-in.

    The intended workflow is: log sessions as the week runs, then fill the weekly
    check-in once it ends. So the newest row that has sessions is usually the week
    still IN PROGRESS — its weight/sleep/stress are not entered yet, which makes its
    status meaningless (it can only ever come out green). "Last completed week"
    therefore means the last week that has a check-in, not the last week touched.
    """
    return f'SUMPRODUCT(MAX((Week!$E${wr1}:$E${wr2}<>"")*ROW(Week!$E${wr1}:$E${wr2})))'


def _since_checkin(col: str, wr1: int, wr2: int, agg: str = "MAX") -> str:
    """Aggregate `col` over the last completed week AND everything after it.

    Used for signals that must not wait for a check-in to surface — finger pain
    logged mid-week has to reach the dashboard on the day it is logged.
    """
    return (f'SUMPRODUCT({agg}((ROW(Week!${col}${wr1}:${col}${wr2})>={_checkin_row(wr1, wr2)})'
            f'*Week!${col}${wr1}:${col}${wr2}))')


def _inprogress(col: str, wr1: int, wr2: int, agg: str = "SUM") -> str:
    """Aggregate `col` strictly AFTER the last check-in — i.e. the week in progress."""
    return (f'SUMPRODUCT({agg}((ROW(Week!${col}${wr1}:${col}${wr2})>{_checkin_row(wr1, wr2)})'
            f'*Week!${col}${wr1}:${col}${wr2}))')


def _phase_today(plan: Plan, wr1: int, wr2: int) -> str:
    """Phase by calendar date (clamped to the plan), independent of check-ins."""
    d = plan.cfg.goal.start_date
    n = len(plan.weeks)
    return (f'=INDEX(Week!$C${wr1}:$C${wr2},'
            f'MIN({n},MAX(1,INT((TODAY()-DATE({d.year},{d.month},{d.day}))/7)+1)))')


def build_dashboard(ws, plan: Plan, wr1: int, wr2: int, jr2: int = 10000) -> None:
    cfg = plan.cfg
    t = translator(cfg.language)

    def q(s, **kw):
        """A translated, Excel-quoted string literal for use inside formulas."""
        return '"' + t(s, **kw).replace('"', '""') + '"'

    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 30, "B": 18, "C": 3, "D": 64})
    ws.merge_cells("A1:D1"); h1(ws["A1"], t("Dashboard — where you are and what to do"))
    ws.row_dimensions[1].height = 26

    unit = t(cfg.profile.units.value)
    start_bw = cfg.profile.bodyweight                 # display units
    target_bw = cfg.weight.target_bodyweight if cfg.weight.enabled else None  # display units
    rate_txt = t("-0.5 kg/wk") if cfg.profile.units.value == "kg" else t("-1.1 lb/wk")
    norm = plan.norm_target_pct
    tgt_v = plan.target_v
    inj = getattr(plan, "_inj_active_formula", '0')

    h2(ws, 3, 1, 2, t("Now (last completed week)"))
    ws.merge_cells("C3:D3"); h2(ws, 3, 3, 4, t("Reading"))

    # (id, name, formula, number_format, description) — id is stable for CF matching
    kpi = [
        ("cur_wt", t("Current weight"), f'=IFERROR({_lastw("E",wr1,wr2)},"—")', "0.0", t("From the latest check-in")),
        ("lost", t("Lost so far"), f'=IFERROR({start_bw}-{_lastw("E",wr1,wr2)},"—")', "0.0", t("From start")),
    ]
    if target_bw is not None:
        kpi.append(("left_to_target", t("Left to target"), f'=IFERROR({_lastw("E",wr1,wr2)}-{target_bw},"—")', "0.0", t("≤0 = weight goal met")))
    kpi += [
        ("phase", t("Current phase"), _phase_today(plan, wr1, wr2), "@", t("Where you are in the plan (by date)")),
        ("fingers", t("Fingers %BW"), f'=IFERROR({_lastne("T",wr1,wr2)},"—")', "0.0%", t("(bw+hang)/bw")),
        ("to_norm", t("To V-target norm"), f'=IFERROR({_lastne("U",wr1,wr2)},"—")', "0.0%", t("≤0 = V{v} finger norm reached", v=tgt_v)),
        ("best_grade", t("Best grade (wk)"), f'=IFERROR({_lastne("V",wr1,wr2)},"—")', "0", t("Max in a week")),
        ("load", t("Week load (sRPE)"), f'=IFERROR({_lastw("Q",wr1,wr2)},"—")', "0", t("Sum of min×RPE")),
        # ramp and monotony are WHOLE-WEEK aggregates: taken from the week in progress
        # they describe a half-finished week and read misleadingly low, so they follow
        # the completed week like load and status do.
        ("acwr", t("Load ramp (vs 4-wk avg)"), f'=IFERROR({_lastw("S",wr1,wr2)},"—")', "0.00", t("~1.0 steady · high = ramped fast (not a risk score)")),
        ("monotony", t("Monotony"), f'=IFERROR({_lastw("AC",wr1,wr2)},"—")', "0.00", t(">2 = every day alike; pair with load")),
        # pain overrides the completed-week status: a mid-week 2+ must not wait for the check-in
        ("status", t("Week status"),
         f'=IFERROR(IF({_since_checkin("W",wr1,wr2)}>=2,"{t("🔴 Finger pain")}",{_lastw("X",wr1,wr2)}),"—")',
         "@", t("Composite traffic light")),
        ("inprog", t("This week so far"),
         f'=IFERROR({_inprogress("P",wr1,wr2)},0)', "0", t("Sessions logged since the last check-in")),
        ("weeks", t("Weeks logged"), f'=COUNT(Week!$E${wr1}:$E${wr2})', "0", t("Check-ins so far")),
        ("sessions", t("Sessions logged"), f'=COUNTIFS(Journal!$D$5:$D${jr2},">0")', "0", t("Journal rows")),
        ("completion", t("Week completion"), f'=IFERROR({_lastw("AB",wr1,wr2)},"—")', "0%", t("Actual ÷ planned sessions")),
        ("injuries", t("Active injuries"), f'={inj}', "0", t("Active + rehab (see Injuries)")),
    ]
    r = 4
    left_to_target_row = acwr_row = inj_row = None
    for kid, name, f, fmt, desc in kpi:
        td(ws.cell(r, 1), name, bold=True)
        ws.cell(r, 2, f); td(ws.cell(r, 2), center=True); ws.cell(r, 2).number_format = fmt
        ws.cell(r, 2).font = Font(FONT, 12, bold=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        td(ws.cell(r, 3), desc)
        if kid == "left_to_target":
            left_to_target_row = r
        elif kid == "acwr":
            acwr_row = r
        elif kid == "injuries":
            inj_row = r
        if (r % 2) == 0:
            for c in (1, 2, 3):
                fill(ws.cell(r, c), GREY)
        ws.row_dimensions[r].height = 26
        r += 1

    if left_to_target_row:
        ws.conditional_formatting.add(f"B{left_to_target_row}", CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor=GREEN)))
    if acwr_row:
        ws.conditional_formatting.add(f"B{acwr_row}", CellIsRule(operator="greaterThan", formula=["1.5"], fill=PatternFill("solid", fgColor=RED)))
    if inj_row:
        ws.conditional_formatting.add(f"B{inj_row}", CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=PatternFill("solid", fgColor=RED)))

    # ---- advisor (concrete actions) ----
    r += 1
    h2(ws, r, 1, 4, t("Advisor")); ws.row_dimensions[r].height = 20
    r += 1
    cw = _lastw("E", wr1, wr2)          # weight KPIs stay anchored on the weigh-in
    g = _lastw("G", wr1, wr2)
    # fingers / gap-to-norm stay on their own column: they are sparse measurements
    # (you don't max-hang every week) and the latest reading is the right one to show.
    # The ramp is a weekly aggregate, so it follows the completed week instead.
    tt, uu = _lastne("T", wr1, wr2), _lastne("U", wr1, wr2)
    ss = _lastw("S", wr1, wr2)
    # recovery / completion / status describe the last COMPLETED week (the one with a
    # check-in). Pain is the exception: it scans from that week onward so anything
    # logged in the week still in progress surfaces immediately.
    yy, zz, ab, xx = (_lastw("Y", wr1, wr2), _lastw("Z", wr1, wr2),
                      _lastw("AB", wr1, wr2), _lastw("X", wr1, wr2))
    ww = _since_checkin("W", wr1, wr2)
    advice = []
    if target_bw is not None:
        hh = _lastw("H", wr1, wr2)          # week-over-week change at the last check-in
        gain_red = 0.5 if cfg.profile.units.value == "kg" else 1.1
        advice.append(
            f'=IFERROR({q("Weight: down ")}&TEXT({start_bw}-{cw},"0.0")&{q(" of ")}&TEXT({start_bw}-{target_bw},"0.0")&{q(" {unit}, ", unit=unit)}&TEXT({cw}-{target_bw},"0.0")&{q(" to go. ")}'
            # a gain while cutting is the sharpest signal — say it before anything else
            f'&IF(AND({hh}<>"",{hh}>{gain_red}),{q("You gained ")}&TEXT({hh},"0.0")&{q(" {unit} this week while cutting — check portions, alcohol and salt/water before changing the plan.", unit=unit)},'
            f'IF({g}>1,{q("Behind the planned curve — add ~100-150 kcal deficit or 1 Zone-2 session.")},'
            f'IF({g}<-1,{q("Faster than planned — raise calories; aim {rate} to keep strength.", rate=rate_txt)},{q("On the planned curve.")}))),{q("Weight: enter a bodyweight in Week.")})'
        )
    advice += [
        f'=IFERROR({q("Fingers: ")}&TEXT({tt},"0.0%")&{q(" BW; to the V{v} norm (", v=tgt_v)}&TEXT({norm},"0%")&{q(") ")}'
        f'&IF({uu}<=0,{q("— reached. Now convert it on the wall (technique, limit).")},'
        f'{q("need ")}&TEXT({uu}*100,"0.0")&{q(" pts (~")}&TEXT({uu}*{cw},"0.0")&{q(" {unit}). Keep 2 finger sessions/wk, add load slowly.", unit=unit)}),{q("Fingers: log a max hang in Journal.")})',
        f'=IFERROR({q("Load ramp: ")}&TEXT({ss},"0.00")&{q(" vs your 4-week average. ")}'
        f'&IF({ss}>1.5,{q("Sharp jump — next week drop 1 power-endurance/volume session and add a rest day.")},'
        f'IF({ss}<0.8,{q("Lighter than your recent average — room to add 1 volume/technique session.")},{q("Close to your recent average — hold steady.")})),{q("Load ramp: needs ~2-3 logged weeks before it means anything.")})',
        f'=IF({inj}>0,{q("⚠ Active injuries: ")}&TEXT({inj},"0")&{q(" — follow the rehab in Injuries, don\'t load the area, swap the affected sessions.")},{q("Injuries: none active.")})',
        f'=IFERROR(IF({ww}>=2,{q("Finger pain ")}&TEXT({ww},"0")&{q("/3 — rest fingers, switch to legs/cardio/mobility, add an Injuries row.")},'
        f'IF(OR(AND({yy}<>"",{yy}<=2),AND({zz}<>"",{zz}>=8)),{q("Sleep/stress are down — swap the next limit day for a technique day, prioritise sleep.")},{q("Recovery: sleep and stress look fine.")})),{q("Log pain, sleep and stress to track recovery.")})',
        f'=IFERROR({q("Completion: ")}&TEXT({ab},"0%")&{q(". ")}&IF({ab}<0.7,{q("Week under-done — don\'t cram it back; add volume gradually (+10-15%).")},{q("Volume on plan.")}),{q("Completion shows once sessions + plan exist.")})',
        f'=IFERROR({q("Overall last-week status: ")}&{xx},{q("Status appears after your first check-in.")})',
    ]
    for a in advice:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1, a); td(ws.cell(r, 1)); fill(ws.cell(r, 1), TEAL)
        ws.row_dimensions[r].height = 34
        r += 1


def build_charts(ws, plan: Plan, wr1: int, wr2: int) -> None:
    import math
    t = translator(plan.cfg.language)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:S1"); h1(ws["A1"], t("Charts (populate as you log)"))
    ws.row_dimensions[1].height = 24
    week = next((s for s in ws.parent.worksheets if s.title == "Week"), None)
    if week is None:
        return
    hdr_row = 4
    cfg = plan.cfg
    m = plan.metrics

    def line(title, cols, anchor, ymin=None, ymax=None, yfmt=None):
        """One line chart; each col in `cols` becomes its own named series.
        ymin/ymax pin the y-axis to the plan's real range (no more 0-based squash)."""
        ch = LineChart(); ch.title = title; ch.height = 7.5; ch.width = 14; ch.style = 2
        for c in cols:
            ref = Reference(week, min_col=c, max_col=c, min_row=hdr_row, max_row=wr2)
            ch.add_data(ref, titles_from_data=True)
        ch.set_categories(Reference(week, min_col=1, min_row=wr1, max_row=wr2))
        ch.x_axis.title = t("Week")
        ch.x_axis.delete = False; ch.y_axis.delete = False
        if ymin is not None:
            ch.y_axis.scaling.min = ymin
        if ymax is not None:
            ch.y_axis.scaling.max = ymax
        if yfmt is not None:
            ch.y_axis.numFmt = yfmt
        ws.add_chart(ch, anchor)

    # ----- data-driven axis bounds (known at generation time) -----
    # Weight: tight band around the planned start→target range, not a 0-based axis.
    start = cfg.profile.bodyweight
    has_cut = cfg.weight.enabled and cfg.weight.target_bodyweight is not None
    target = cfg.weight.target_bodyweight if has_cut else start
    w_lo, w_hi = min(start, target), max(start, target)
    w_min, w_max = math.floor(w_lo - 4), math.ceil(w_hi + 4)

    # Fingers %BW: focus on [current strength .. target norm] with a little padding.
    norm = plan.norm_target_pct
    cur_pct = (cfg.bw_kg + cfg.max_hang_kg) / cfg.bw_kg if cfg.max_hang_kg is not None else None
    f_lo = min(cur_pct, norm) if cur_pct is not None else norm - 0.4
    f_min = max(0.5, math.floor((f_lo - 0.1) * 10) / 10)
    f_max = math.ceil((norm + 0.1) * 10) / 10

    # Grade: around current → target.
    g_min, g_max = max(0, plan.current_v - 2), plan.target_v + 2
    # Sessions: 0 → a bit above the busiest planned week.
    s_max = max((w.planned_session_count for w in plan.weeks), default=6) + 1

    # ----- charts (2-up: left column A, right column K) -----
    line(t("Weight: actual vs plan") if has_cut else t("Weight"),
         [5, 6] if has_cut else [5], "A3", w_min, w_max)
    line(t("Fingers %BW (target {pct}%)", pct=f"{norm*100:.0f}"), [20], "K3", f_min, f_max, yfmt="0%")

    line(t("Best grade / week (V)"), [22], "A18", g_min, g_max)
    line(t("Weekly load: acute vs chronic (sRPE)"), [17, 18], "K18", 0)

    line(t("Load ramp vs 4-week average"), [19], "A33", 0, 2.5)
    line(t("Sessions: planned vs done"), [27, 16], "K33", 0, s_max)

    line(t("Fatigue & stress (1-10)"), [15, 26], "A48", 0, 10)
    line(t("Sleep (h / night)"), [14], "K48", 0, 12)

    line(t("Max finger pain (0-3)"), [23], "A63", 0, 3)
    if m["hrv"] or m["resting_hr"]:
        hrv_cols = ([10] if m["hrv"] else []) + ([12] if m["resting_hr"] else [])
        line(t("HRV & resting HR"), hrv_cols, "K63")   # auto y-axis (device-dependent range)

    # Foster's pair: sameness of training, and load weighted by that sameness
    line(t("Monotony (day-to-day sameness)"), [29], "A78", 0, 4)
    line(t("Strain (load × monotony)"), [30], "K78", 0)

    note(ws.cell(95, 1),
         t("Lines appear as Week fills in. Axes are pre-scaled to your plan's range. The load ramp "
           "and monotony describe what changed — they are not injury predictions; see the Glossary."))
    ws.merge_cells("A95:S95")
