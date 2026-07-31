"""
crimped.render.style — shared styling helpers for all sheets.

Centralizes the palette, fonts and cell helpers so every sheet looks consistent
and the sheet modules stay focused on content/formulas.
"""

from __future__ import annotations
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "Arial"

# palette — mirrors the web app's light theme ("sandstone + alpenglow"):
# ink #1E2732, alpine blue #2F7EA6, alpenglow orange #D9601F, warm greys.
# Legacy names are kept so sheet modules don't churn; only the values changed.
NAVY = "1E2732"      # ink — h1 banner
BLUE = "2F7EA6"      # alpine — h2 / table headers
LBLUE = "DCEAF2"     # light alpine — computed highlights
GREEN = "D9E6CF"     # soft moss — good / deficit weeks
YELLOW = "FCF0DC"    # warm chalk — user-input cells
ORANGE = "F8E3D2"    # soft alpenglow — deloads / cautions
RED = "F3CFC5"       # soft clay — red flags
GREY = "EFECE6"      # sandstone — row banding
TEAL = "FDEBDC"      # pale alpenglow — advisor rows (the "act on this" area)
GLOW = "D9601F"      # alpenglow accent — chart lines, emphasis
ALPINE = "2F7EA6"    # chart secondary series

_thin = Side(style="thin", color="DDD7CC")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def solid(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def h1(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(FONT, 15, bold=True, color="FFFFFF")
    cell.fill = solid(NAVY)
    cell.alignment = Alignment("left", "center")


def h2(ws, row: int, c1: int, c2: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row, c1, text)
    cell.font = Font(FONT, 12, bold=True, color="FFFFFF")
    cell.fill = solid(BLUE)
    cell.alignment = Alignment("left", "center")


def th(cell, text=None) -> None:
    if text is not None:
        cell.value = text
    cell.font = Font(FONT, 9, bold=True, color="FFFFFF")
    cell.fill = solid(BLUE)
    cell.alignment = Alignment("center", "center", wrap_text=True)
    cell.border = BORDER


def td(cell, text=None, center=False, bold=False, size=10) -> None:
    if text is not None:
        cell.value = text
    cell.font = Font(FONT, size, bold=bold)
    cell.border = BORDER
    cell.alignment = Alignment("center" if center else "left", "top", wrap_text=True)


def fill(cell, color: str) -> None:
    cell.fill = solid(color)


def inp(cell) -> None:
    """Mark a cell as user-input (yellow)."""
    cell.fill = solid(YELLOW)
    cell.border = BORDER
    cell.alignment = Alignment("center", "center")


def note(cell, text=None) -> None:
    if text is not None:
        cell.value = text
    cell.font = Font(FONT, 10, italic=True, color="6B7A89")   # web --muted
    cell.alignment = Alignment("left", "top", wrap_text=True)


def widths(ws, mapping: dict) -> None:
    """mapping: {'A': 12, 'B': 30, ...}"""
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w
